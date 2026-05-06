from cmdbox.app import filer
from cmdbox.app.commons import convert, resdata, redis_client
from cmdbox.app.features.cli.excel import excel_base
from pathlib import Path
from typing import Dict, Any, List, Union
import argparse
import logging
import pydantic


class ExcelSheetList(excel_base.ExcelBase):
    def get_cmd(self):
        """
        この機能のコマンドを返します

        Returns:
            str: コマンド
        """
        return 'sheet_list'
    
    def get_option(self):
        """
        この機能のオプションを返します

        Returns:
            Dict[str, Any]: オプション
        """
        opt = super().get_option()
        opt['description_ja'] = "データフォルダ配下のExcelファイルのシート一覧を取得します。"
        opt['description_en'] = "Retrieves the list of sheets in an Excel file located within the data folder."
        return opt

    def excel_proc(self, abspath:Path, args:argparse.Namespace, logger:logging.Logger, tm:float, pf:List[Dict[str, float]]=[]) -> Dict[str, Any]:
        """
        Excel処理のベース

        Args:
            abspath (Path): Excelファイルの絶対パス
            args (argparse.Namespace): 引数
            logger (logging.Logger): ロガー
            tm (float): 処理時間
            pf (List[Dict[str, float]]): パフォーマンス情報

        Returns:
            Dict[str, Any]: 結果
        """
        res_json = self.sheet_list(abspath, logger)
        return res_json

    def get_svparam(self, args:argparse.Namespace) -> List[str]:
        """
        サーバーに送信するパラメーターを返します

        Args:
            args (argparse.Namespace): 引数

        Returns:
            List[str]: サーバーに送信するパラメーター
        """
        ret = [convert.str2b64str(str(args.svpath))]
        return ret

    def output_schema(self) -> type:
        class Row(resdata.Base):
            sheet_name: str = pydantic.Field(description="シート名")
            sheetinfos: Union[Dict[str, Any], str, None] = pydantic.Field(default=None, description="シート情報")
        class Data(resdata.Data):
            data: Union[List[Row], None] = pydantic.Field(default=None, description="処理結果のデータ")
        class Result(resdata.Result):
            success: Union[Data, None] = pydantic.Field(default=None, description="成功した場合の結果")
        return Result

    def is_cluster_redirect(self):
        """
        クラスター宛のメッセージの場合、メッセージを転送するかどうかを返します

        Returns:
            bool: メッセージを転送する場合はTrue
        """
        return False

    def svrun(self, data_dir:Path, logger:logging.Logger, redis_cli:redis_client.RedisClient, msg:List[str],
              sessions:Dict[str, Dict[str, Any]]) -> int:
        """
        この機能のサーバー側の実行を行います

        Args:
            data_dir (Path): データディレクトリ
            logger (logging.Logger): ロガー
            redis_cli (redis_client.RedisClient): Redisクライアント
            msg (List[str]): 受信メッセージ
            sessions (Dict[str, Dict[str, Any]]): セッション情報
        
        Returns:
            int: 終了コード
        """
        svpath = convert.b64str2str(msg[2])

        try:
            f = filer.Filer(data_dir, logger)
            chk, abspath, res = f._file_exists(svpath)
            if not chk:
                logger.warning(f"File not found. {svpath}")
                redis_cli.rpush(msg[1], res)
                return self.RESP_WARN
            res = self.sheet_list(abspath, logger)
            redis_cli.rpush(msg[1], res)
        except Exception as e:
            logger.warning(f"Failed to sheet list: {e}", exc_info=True)
            redis_cli.rpush(msg[1], dict(warn=f"Failed to sheet list: {e}"))
            return self.RESP_WARN
        return self.RESP_SUCCESS

    def sheet_list(self, filepath:str, logger:logging.Logger) -> Dict[str, Any]:
        """
        指定したワークブックのシート一覧を取得します。

        Args:
            filepath (str): ワークブックのパス
            logger (logging.Logger): ロガー
        Returns:
            dict: シートの情報
        """
        wb:Workbook = None
        try:
            from openpyxl.cell import Cell
            from openpyxl.workbook.workbook import Workbook
            from openpyxl.worksheet.worksheet import Worksheet
            import openpyxl

            wb:Workbook = openpyxl.load_workbook(filename=filepath, read_only=True)
            rows = []
            for sheet in wb.worksheets:
                rows.append(dict(sheet_name=sheet.title, sheetinfos=dict(max_row=sheet.max_row, max_column=sheet.max_column, sheet_state=sheet.sheet_state)))

            res = dict(success=dict(data=rows))
            return res
        except Exception as e:
            msg = dict(warn=f"Failed to sheet list: {e}")
            logger.warning(f"Failed to sheet list: {e}", exc_info=True)
            return msg
        finally:
            if wb is not None:
                wb.close()
